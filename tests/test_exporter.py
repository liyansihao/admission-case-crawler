from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from admission_case_crawler.config import load_config
from admission_case_crawler.default_config import DEFAULT_CONFIG
from admission_case_crawler import exporter
from admission_case_crawler.exporter import build_outputs


def test_build_outputs_merges_platforms_and_hides_false(tmp_path: Path) -> None:
    (tmp_path / "config.yaml").write_text(DEFAULT_CONFIG, encoding="utf-8")
    raw = tmp_path / "raw"
    for platform in ("xhs", "weibo", "wechat"):
        path = raw / platform / "jsonl"
        path.mkdir(parents=True)
        record = {
            "note_id": f"{platform}-1",
            "title": "港硕录取 offer",
            "desc": "本科 GPA 3.7/4.0 香港大学 offer",
            "nickname": f"{platform}-account",
            "time": 1782885600,
            "url": f"https://example.com/{platform}",
        }
        (path / "search_contents_2026-07-03.jsonl").write_text(json.dumps(record, ensure_ascii=False) + "\n", encoding="utf-8")

    cfg = load_config(tmp_path / "config.yaml")
    records = build_outputs(cfg)

    assert len(records) == 3
    workbook = openpyxl.load_workbook(tmp_path / "output" / "excel" / "微博小红书公众号录取案例汇总_无False.xlsx", read_only=False, data_only=True)
    assert workbook.sheetnames == ["逐条案例", "汇总图图片提取", "图片页索引", "候选但未细读", "搜索关键词与来源记录"]
    sheet = workbook["逐条案例"]
    assert [sheet.cell(1, column).value for column in range(1, 6)] == ["来源平台", "搜索关键词", "笔记标题", "笔记 ID", "链接"]
    assert sheet.auto_filter.ref == "A1:AA4"
    assert str(sheet.freeze_panes) == "A2"
    values = [cell for sheet_name in workbook.sheetnames for row in workbook[sheet_name].iter_rows(values_only=True) for cell in row]
    assert False not in values
    assert "false" not in [str(value).lower() for value in values if value is not None]


def test_build_outputs_uses_ocr_text_for_extraction(tmp_path: Path, monkeypatch) -> None:
    config_text = DEFAULT_CONFIG.replace("enabled: false", "enabled: true", 1)
    (tmp_path / "config.yaml").write_text(config_text, encoding="utf-8")
    raw_dir = tmp_path / "raw" / "xhs" / "jsonl"
    raw_dir.mkdir(parents=True)
    record = {
        "note_id": "xhs-image-1",
        "title": "图片案例",
        "desc": "正文没有结构化字段",
        "nickname": "xhs-account",
        "time": 1782885600,
        "url": "https://example.com/xhs-image-1",
        "image_list": ["https://example.com/image.jpg"],
    }
    (raw_dir / "search_contents_2026-07-03.jsonl").write_text(json.dumps(record, ensure_ascii=False) + "\n", encoding="utf-8")

    def fake_enrich_raw_with_ocr(cfg, platform, note_id, raw):
        raw = dict(raw)
        raw["ocr_text"] = "香港大学 offer GPA 3.8/4.0"
        raw["ocr_image_paths"] = [str(tmp_path / "raw" / "images" / "xhs-image-1.jpg")]
        return raw

    monkeypatch.setattr(exporter, "enrich_raw_with_ocr", fake_enrich_raw_with_ocr)

    cfg = load_config(tmp_path / "config.yaml")
    records = build_outputs(cfg)

    assert len(records) == 1
    assert records[0]["from_image"] is True
    assert records[0]["application_school"] == "香港大学"
    assert records[0]["admission_result"] == "offer"

    workbook = openpyxl.load_workbook(tmp_path / "output" / "excel" / "微博小红书公众号录取案例汇总_无False.xlsx", data_only=True)
    detail_values = [cell for row in workbook["逐条案例"].iter_rows(values_only=True) for cell in row if cell]
    image_values = [cell for row in workbook["图片页索引"].iter_rows(values_only=True) for cell in row if cell]
    assert any("图片 OCR" in str(value) for value in detail_values)
    assert any("已 OCR" in str(value) for value in image_values)
