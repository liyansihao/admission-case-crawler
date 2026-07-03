from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .config import AppConfig
from .ocr import enrich_raw_with_ocr


DETAIL_HEADERS = [
    "来源平台",
    "搜索关键词",
    "笔记标题",
    "笔记 ID",
    "链接",
    "作者/账号名",
    "发布时间/编辑时间",
    "录取学校",
    "项目方向",
    "项目英文名",
    "学位层次/项目类型",
    "录取年份",
    "Offer 时间",
    "录取结果",
    "本科院校背景",
    "本科专业",
    "GPA / 均分 / 等级",
    "雅思 / 托福",
    "GMAT / GRE",
    "实习经历",
    "科研 / 竞赛 / 奖学金",
    "工作背景",
    "申请时间线",
    "面试 / VI 题目",
    "其他补充",
    "信息来源类型",
    "可信度或备注",
]

SUMMARY_HEADERS = [
    "来源平台",
    "搜索关键词",
    "笔记标题",
    "笔记 ID",
    "链接",
    "作者/账号名",
    "发布时间/编辑时间",
    "学校",
    "提取对象",
    "项目方向",
    "项目英文名",
    "录取年份",
    "提取内容",
    "涉及 BG / 成绩 / 语言",
    "信息来源类型",
    "可信度或备注",
]

IMAGE_HEADERS = [
    "来源平台",
    "搜索关键词",
    "笔记标题",
    "笔记 ID",
    "链接",
    "作者/账号名",
    "发布时间/编辑时间",
    "图片文件",
    "本地截图路径",
    "图片类型",
    "处理状态",
    "备注",
]

CANDIDATE_HEADERS = [
    "来源平台",
    "搜索关键词",
    "笔记标题",
    "笔记 ID",
    "链接",
    "作者/账号名",
    "发布时间/编辑时间",
    "处理状态",
    "筛选备注",
]

SOURCE_HEADERS = [
    "项目",
    "内容",
    "备注",
]

REGION_TERMS = {
    "香港": ["香港", "港硕", "港校", "HKU", "CUHK", "HKUST", "CityU", "PolyU"],
    "英国": ["英国", "英硕", "G5", "KCL", "UCL", "Manchester", "Edinburgh"],
    "美国": ["美国", "美研", "US", "USA"],
    "新加坡": ["新加坡", "NUS", "NTU"],
    "澳洲": ["澳洲", "澳大利亚", "Melbourne", "Sydney"],
}

SCHOOL_TERMS = {
    "香港大学": ["香港大学", "港大", "HKU"],
    "香港中文大学": ["香港中文大学", "港中文", "CUHK"],
    "香港科技大学": ["香港科技大学", "港科", "HKUST"],
    "香港城市大学": ["香港城市大学", "港城", "CityU"],
    "新加坡国立大学": ["新加坡国立大学", "NUS"],
    "南洋理工大学": ["南洋理工", "NTU"],
    "伦敦大学学院": ["伦敦大学学院", "UCL"],
    "伦敦国王学院": ["伦敦国王学院", "KCL", "King's College London"],
}


def build_outputs(cfg: AppConfig) -> list[dict[str, Any]]:
    comments = load_comments(cfg.raw_dir)
    records = []
    for platform, raw in load_content_records(cfg.raw_dir):
        note_id = text(raw.get("note_id") or raw.get("id") or "")
        raw = enrich_raw_with_ocr(cfg, platform, note_id, raw)
        records.append(normalize_record(platform, raw, comments.get(f"{platform}:{note_id}", [])))
    records = dedupe(records)
    mark_incremental(records, cfg.state_file)
    write_outputs(records, cfg.output_dir)
    return records


def load_content_records(raw_dir: Path) -> Iterable[tuple[str, dict[str, Any]]]:
    if not raw_dir.exists():
        return
    for platform_dir in sorted(raw_dir.iterdir()):
        jsonl_dir = platform_dir / "jsonl"
        if not jsonl_dir.exists():
            continue
        for jsonl in sorted(jsonl_dir.glob("*_contents_*.jsonl")):
            for item in read_jsonl(jsonl):
                yield platform_dir.name, item


def load_comments(raw_dir: Path) -> dict[str, list[dict[str, Any]]]:
    by_note: dict[str, list[dict[str, Any]]] = {}
    if not raw_dir.exists():
        return by_note
    for platform_dir in sorted(raw_dir.iterdir()):
        jsonl_dir = platform_dir / "jsonl"
        if not jsonl_dir.exists():
            continue
        for jsonl in sorted(jsonl_dir.glob("*_comments_*.jsonl")):
            for item in read_jsonl(jsonl):
                note_id = text(item.get("note_id"))
                if note_id:
                    by_note.setdefault(f"{platform_dir.name}:{note_id}", []).append(item)
    return by_note


def read_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(value, dict):
                yield value


def normalize_record(platform: str, raw: dict[str, Any], comments: list[dict[str, Any]]) -> dict[str, Any]:
    note_id = text(raw.get("note_id") or raw.get("id") or "")
    title = text(raw.get("title") or "")
    body = text(raw.get("desc") or raw.get("content") or raw.get("raw_text") or "")
    raw_text = combine(title, body)
    ocr_text = text(raw.get("ocr_text") or "")
    comments_text = "\n".join(text(item.get("content")) for item in comments if item.get("content"))
    full_text = "\n".join(part for part in [raw_text, comments_text, ocr_text] if part)
    extracted = extract_fields(full_text)
    source_id = f"{platform}:{note_id}" if note_id else f"{platform}:{stable_hash(raw_text)}"
    content_type = classify(full_text)
    needs_review = not bool(raw_text and (raw.get("note_url") or raw.get("url"))) or content_type == "other"
    return {
        "platform": "weibo" if platform == "wb" else platform,
        "source_platform": platform_label(platform),
        "search_keyword": text(raw.get("source_keyword") or raw.get("keyword") or raw.get("source_account_keyword") or ""),
        "note_id": note_id,
        "source_account": text(raw.get("nickname") or raw.get("source_account") or raw.get("author_name") or ""),
        "publish_time": normalize_time(raw.get("time") or raw.get("create_time") or raw.get("create_date_time")),
        "url": text(raw.get("note_url") or raw.get("url") or ""),
        "content_type": content_type,
        "raw_text": raw_text,
        "comments_text": comments_text,
        "ocr_text": ocr_text,
        "undergraduate_school": extracted["undergraduate_school"],
        "undergraduate_major": extracted["undergraduate_major"],
        "gpa": extracted["gpa"],
        "gpa_scale": extracted["gpa_scale"],
        "toefl_ielts": extracted["toefl_ielts"],
        "gre_gmat": extracted["gre_gmat"],
        "application_region": extracted["application_region"],
        "application_school": extracted["application_school"],
        "application_program": extracted["application_program"],
        "admission_result": extracted["admission_result"],
        "admission_year": extracted["admission_year"],
        "is_repost": bool(raw.get("retweeted_status") or "//@" in full_text or "转发" in full_text),
        "from_image": bool(ocr_text),
        "confidence": confidence(extracted, content_type),
        "needs_review": needs_review,
        "is_new": True,
        "notes": "",
        "source_id": source_id,
        "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "title": title or first_line(raw_text) or note_id,
        "project_english_name": extract_project_english(full_text),
        "degree_type": infer_degree_type(full_text),
        "offer_time": first_match(full_text, [r"(?:Offer|offer|录取)(?:时间|日期)?[:： ]{0,3}([0-9]{4}[-/.][0-9]{1,2}[-/.][0-9]{1,2}|[0-9]{1,2}月[0-9]{1,2}日)"]),
        "internship": first_match(full_text, [r"(?:实习|实习经历)[:： ]{0,3}([^。；;\n]{2,80})"]),
        "research_competition_scholarship": first_match(full_text, [r"(?:科研|竞赛|奖学金)[:： ]{0,3}([^。；;\n]{2,80})"]),
        "work_background": first_match(full_text, [r"(?:工作|工作背景)[:： ]{0,3}([^。；;\n]{2,80})"]),
        "application_timeline": first_match(full_text, [r"(?:时间线|申请时间线)[:： ]{0,3}([^。；;\n]{2,100})"]),
        "interview_vi": first_match(full_text, [r"(?:面试|VI)[:： ]{0,3}([^。；;\n]{2,100})"]),
        "image_urls": parse_url_list(raw.get("image_list") or raw.get("pictures") or raw.get("cover") or ""),
        "ocr_image_paths": parse_url_list(raw.get("ocr_image_paths") or ""),
    }


def extract_fields(value: str) -> dict[str, str]:
    return {
        "undergraduate_school": first_match(value, [r"(?:本科|本科学校|院校)[:： ]{0,3}([^,，。；;\n]{2,30})"]),
        "undergraduate_major": first_match(value, [r"(?:专业|本科专业)[:： ]{0,3}([^,，。；;\n]{2,30})"]),
        "gpa": first_match(value, [r"GPA[:： ]?([0-9.]{1,4})", r"均分[:： ]?([0-9.]{2,5})"]),
        "gpa_scale": first_match(value, [r"GPA[:： ]?[0-9.]{1,4}\s*/\s*([0-9.]{1,4})"]),
        "toefl_ielts": first_match(value, [r"(?:雅思|IELTS)[:： ]?([0-9.]{1,4})", r"(?:托福|TOEFL)[:： ]?([0-9]{2,3})"]),
        "gre_gmat": first_match(value, [r"(?:GRE|GMAT)[:： ]?([0-9]{2,3})"]),
        "application_region": alias_lookup(value, REGION_TERMS),
        "application_school": alias_lookup(value, SCHOOL_TERMS),
        "application_program": first_match(value, [r"(?:项目|专业|programme|program)[:： ]{0,3}([^,，。；;\n]{2,50})"]),
        "admission_result": result(value),
        "admission_year": first_match(value, [r"(20[0-9]{2})(?:fall|Fall|秋|申请|录取)?"]),
    }


def extract_project_english(value: str) -> str:
    patterns = [
        r"\b((?:MSc|MA|MSSc|MPhil|Master|MBA|MPA|LLM|PhD)[A-Za-z\s&(),/-]{4,90})",
        r"\b(Bachelor [A-Za-z\s&(),/-]{4,90})",
    ]
    result_value = first_match(value, patterns)
    return result_value.strip(" .;；,，")


def infer_degree_type(value: str) -> str:
    if re.search(r"博士|PhD|DPhil", value, re.I):
        return "博士"
    if re.search(r"硕士|研究生|MSc|MA|MSSc|MPhil|Master|MBA|MPA|LLM", value, re.I):
        return "硕士"
    if re.search(r"本科|Bachelor|BSc|BA", value, re.I):
        return "本科"
    return ""


def classify(value: str) -> str:
    if re.search(r"offer|录取|admit|admitted", value, re.I):
        return "admission_case"
    if re.search(r"面经|面试", value):
        return "interview"
    if re.search(r"申请经验|申请复盘|时间线|文书", value):
        return "application_experience"
    return "other"


def result(value: str) -> str:
    if re.search(r"waitlist|wl|候补", value, re.I):
        return "waitlist"
    if re.search(r"拒信|被拒|reject|rejection", value, re.I):
        return "reject"
    if re.search(r"offer|录取|admit|admitted", value, re.I):
        return "offer"
    return ""


def confidence(extracted: dict[str, str], content_type: str) -> str:
    score = sum(bool(extracted.get(key)) for key in ["gpa", "application_school", "admission_result"])
    if content_type == "other":
        return "low"
    return "high" if score >= 2 else "medium"


def dedupe(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    unique: list[dict[str, Any]] = []
    for record in records:
        key = text(record.get("source_id")) or stable_hash(record.get("raw_text", ""))
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)
    return unique


def mark_incremental(records: list[dict[str, Any]], state_file: Path) -> None:
    previous: set[str] = set()
    if state_file.exists():
        try:
            previous = set(json.loads(state_file.read_text(encoding="utf-8")).get("seen_source_ids", []))
        except json.JSONDecodeError:
            previous = set()
    current = set(previous)
    for record in records:
        source_id = text(record.get("source_id"))
        record["is_new"] = source_id not in previous
        if source_id:
            current.add(source_id)
    state_file.parent.mkdir(parents=True, exist_ok=True)
    state_file.write_text(json.dumps({"seen_source_ids": sorted(current)}, ensure_ascii=False, indent=2), encoding="utf-8")


def write_outputs(records: list[dict[str, Any]], output_dir: Path) -> None:
    excel_dir = output_dir / "excel"
    markdown_dir = output_dir / "markdown"
    excel_dir.mkdir(parents=True, exist_ok=True)
    markdown_dir.mkdir(parents=True, exist_ok=True)

    main_xlsx = excel_dir / "admission_cases.xlsx"
    cn_xlsx = excel_dir / "微博小红书公众号录取案例汇总_无False.xlsx"
    write_reference_format_workbook(records, main_xlsx)
    write_reference_format_workbook(records, cn_xlsx)

    for record in records:
        filename = f"{safe_filename(record.get('publish_time') or 'unknown')}_{safe_filename(record.get('source_id'))}.md"
        (markdown_dir / filename).write_text(render_markdown(record), encoding="utf-8")

    summary = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "record_count": len(records),
        "new_count": sum(1 for item in records if item.get("is_new")),
        "needs_review_count": sum(1 for item in records if item.get("needs_review")),
    }
    (output_dir / "update_log.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def write_reference_format_workbook(records: list[dict[str, Any]], path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    detail_rows = [detail_row(record) for record in records if record.get("content_type") == "admission_case"]
    candidate_rows = [candidate_row(record) for record in records if record.get("content_type") != "admission_case" or record.get("needs_review")]
    summary_rows = [summary_row(record) for record in records if record.get("content_type") == "admission_case" and record.get("needs_review")]
    image_rows = image_index_rows(records)
    source_rows = source_log_rows(records)

    add_sheet(workbook, "逐条案例", DETAIL_HEADERS, detail_rows, detail_widths())
    add_sheet(workbook, "汇总图图片提取", SUMMARY_HEADERS, summary_rows, summary_widths())
    add_sheet(workbook, "图片页索引", IMAGE_HEADERS, image_rows, image_widths())
    add_sheet(workbook, "候选但未细读", CANDIDATE_HEADERS, candidate_rows, candidate_widths())
    add_sheet(workbook, "搜索关键词与来源记录", SOURCE_HEADERS, source_rows, source_widths())
    workbook.save(path)


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]], widths: list[float]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([excel_value(value) for value in row])
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def detail_row(record: dict[str, Any]) -> list[Any]:
    return [
        record.get("source_platform", ""),
        record.get("search_keyword", ""),
        record.get("title", ""),
        record.get("note_id", ""),
        record.get("url", ""),
        record.get("source_account", ""),
        record.get("publish_time", ""),
        record.get("application_school", ""),
        record.get("application_program", ""),
        record.get("project_english_name", ""),
        record.get("degree_type", ""),
        format_admission_year(record.get("admission_year", "")),
        record.get("offer_time", ""),
        result_label(record.get("admission_result", "")),
        record.get("undergraduate_school", ""),
        record.get("undergraduate_major", ""),
        join_nonempty([record.get("gpa", ""), record.get("gpa_scale", "")], " / "),
        record.get("toefl_ielts", ""),
        record.get("gre_gmat", ""),
        record.get("internship", ""),
        record.get("research_competition_scholarship", ""),
        record.get("work_background", ""),
        record.get("application_timeline", ""),
        record.get("interview_vi", ""),
        other_supplement(record),
        source_type(record),
        credibility_note(record),
    ]


def summary_row(record: dict[str, Any]) -> list[Any]:
    return [
        record.get("source_platform", ""),
        record.get("search_keyword", ""),
        record.get("title", ""),
        record.get("note_id", ""),
        record.get("url", ""),
        record.get("source_account", ""),
        record.get("publish_time", ""),
        record.get("application_school", ""),
        "正文/候选汇总",
        record.get("application_program", ""),
        record.get("project_english_name", ""),
        format_admission_year(record.get("admission_year", "")),
        trim(record.get("raw_text", ""), 180),
        join_nonempty([record.get("undergraduate_school", ""), record.get("gpa", ""), record.get("toefl_ielts", "")], "；"),
        source_type(record),
        credibility_note(record),
    ]


def candidate_row(record: dict[str, Any]) -> list[Any]:
    return [
        record.get("source_platform", ""),
        record.get("search_keyword", ""),
        record.get("title", ""),
        record.get("note_id", ""),
        record.get("url", ""),
        record.get("source_account", ""),
        record.get("publish_time", ""),
        "候选但未细读" if record.get("needs_review") else "非逐条案例",
        "标题/正文相关，但字段不完整，需人工复核" if record.get("needs_review") else "未识别为录取案例",
    ]


def image_index_rows(records: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for record in records:
        local_paths = record.get("ocr_image_paths") or []
        for index, image_url in enumerate(record.get("image_urls") or [], start=1):
            local_path = local_paths[index - 1] if index - 1 < len(local_paths) else image_url
            rows.append(
                [
                    record.get("source_platform", ""),
                    record.get("search_keyword", ""),
                    record.get("title", ""),
                    record.get("note_id", ""),
                    record.get("url", ""),
                    record.get("source_account", ""),
                    record.get("publish_time", ""),
                    f"image_{index}",
                    local_path,
                    "原始图片/封面",
                    "已 OCR" if record.get("ocr_text") else "已记录图片链接",
                    trim(record.get("ocr_text", ""), 160) if record.get("ocr_text") else "如字段未完整拆分，可按图片链接复核原图",
                ]
            )
    return rows


def source_log_rows(records: list[dict[str, Any]]) -> list[list[Any]]:
    platforms = sorted({record.get("source_platform", "") for record in records if record.get("source_platform")})
    keywords = sorted({record.get("search_keyword", "") for record in records if record.get("search_keyword")})
    return [
        ["抓取平台", "、".join(platforms), "由微博、小红书、公众号原始数据合并"],
        ["搜索关键词", "、".join(keywords), "为空表示原始平台未返回关键词字段"],
        ["读取候选数", str(len(records)), "去重后的记录数"],
        ["逐条案例数", str(sum(1 for item in records if item.get("content_type") == "admission_case")), "进入主表"],
        ["候选但未细读数", str(sum(1 for item in records if item.get("content_type") != "admission_case" or item.get("needs_review"))), "需要人工复核"],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "本地生成"],
    ]


def render_markdown(record: dict[str, Any]) -> str:
    lines = [
        f"# {record.get('title') or record.get('source_id')}",
        "",
        f"- platform: {record.get('platform', '')}",
        f"- account: {record.get('source_account', '')}",
        f"- publish_time: {record.get('publish_time', '')}",
        f"- url: {record.get('url', '')}",
        f"- result: {record.get('admission_result', '')}",
        f"- school: {record.get('application_school', '')}",
        "",
        "## Raw Text",
        "",
        str(record.get("raw_text", "")),
    ]
    if record.get("ocr_text"):
        lines.extend(["", "## OCR Text", "", str(record.get("ocr_text", ""))])
    return "\n".join(lines).strip() + "\n"


def combine(title: Any, body: Any) -> str:
    parts = [text(title), text(body)]
    parts = [part for part in parts if part]
    if len(parts) == 2 and parts[0] == parts[1]:
        return parts[0]
    return "\n\n".join(parts)


def normalize_time(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, str) and not value.isdigit():
        return text(value)
    try:
        number = int(value)
    except (TypeError, ValueError):
        return text(value)
    if number > 10_000_000_000:
        number //= 1000
    return datetime.fromtimestamp(number).strftime("%Y-%m-%d %H:%M:%S")


def platform_label(platform: str) -> str:
    normalized = "weibo" if platform == "wb" else platform
    return {
        "xhs": "小红书",
        "weibo": "微博",
        "wechat": "公众号",
    }.get(normalized, normalized)


def first_match(value: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, value, re.I)
        if match:
            return text(match.group(1))
    return ""


def alias_lookup(value: str, aliases: dict[str, list[str]]) -> str:
    lower = value.lower()
    for label, terms in aliases.items():
        if any(term.lower() in lower for term in terms):
            return label
    return ""


def excel_value(value: Any) -> Any:
    if isinstance(value, bool):
        return "是" if value else ""
    return value


def result_label(value: Any) -> str:
    value = text(value).lower()
    if value == "offer":
        return "录取/offer"
    if value == "reject":
        return "拒信/reject"
    if value == "waitlist":
        return "waitlist"
    return value


def format_admission_year(value: Any) -> str:
    value = text(value)
    if re.fullmatch(r"20[0-9]{2}", value):
        return value[-2:] + "Fall"
    return value


def source_type(record: dict[str, Any]) -> str:
    if record.get("platform") == "wechat":
        return "公众号文章"
    if record.get("ocr_text") and record.get("comments_text"):
        return "正文 + 评论 + 图片 OCR"
    if record.get("ocr_text"):
        return "正文 + 图片 OCR"
    if record.get("comments_text"):
        return "正文 + 评论"
    return "正文"


def credibility_note(record: dict[str, Any]) -> str:
    confidence_value = record.get("confidence", "")
    if record.get("needs_review"):
        return "候选价值高，需人工复核"
    if confidence_value == "high":
        return "正文明确"
    if confidence_value == "medium":
        return "字段部分明确"
    return "需人工复核"


def other_supplement(record: dict[str, Any]) -> str:
    parts = []
    if record.get("application_region"):
        parts.append(f"申请地区：{record['application_region']}")
    if record.get("comments_text"):
        parts.append("含评论信息")
    if record.get("ocr_text"):
        parts.append("图片 OCR：" + trim(record.get("ocr_text", ""), 180))
    raw = trim(record.get("raw_text", ""), 120)
    if raw:
        parts.append(raw)
    return "；".join(parts)


def join_nonempty(values: list[Any], sep: str = "；") -> str:
    return sep.join(text(value) for value in values if text(value))


def trim(value: Any, limit: int) -> str:
    value = text(value)
    if len(value) <= limit:
        return value
    return value[:limit].rstrip() + "..."


def parse_url_list(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [text(item) for item in value if text(item)]
    if isinstance(value, dict):
        return [text(item) for item in value.values() if text(item)]
    raw = text(value)
    try:
        parsed = json.loads(raw)
        return parse_url_list(parsed)
    except (json.JSONDecodeError, TypeError):
        pass
    if raw.startswith("http"):
        return [raw]
    return [part.strip() for part in re.split(r"[,，\n]", raw) if part.strip()]


def detail_widths() -> list[float]:
    return [18, 18, 28, 18, 36, 18, 18, 18, 18, 28, 18, 18, 18, 18, 28, 18, 18, 18, 18, 36, 18, 18, 36, 36, 36, 18, 36]


def summary_widths() -> list[float]:
    return [18, 18, 28, 18, 36, 18, 18, 18, 18, 18, 28, 18, 36, 36, 18, 36]


def image_widths() -> list[float]:
    return [18, 18, 28, 18, 36, 18, 18, 18, 36, 18, 18, 36]


def candidate_widths() -> list[float]:
    return [18, 18, 28, 18, 36, 18, 18, 18, 36]


def source_widths() -> list[float]:
    return [18, 36, 36]


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def first_line(value: str) -> str:
    return value.splitlines()[0].strip() if value else ""


def stable_hash(value: Any) -> str:
    return hashlib.sha1(text(value).encode("utf-8")).hexdigest()[:16]


def safe_filename(value: Any) -> str:
    value = text(value).replace(":", "-")
    value = re.sub(r'[<>"/\\|?*\s]+', "_", value)
    return value[:120] or "untitled"
